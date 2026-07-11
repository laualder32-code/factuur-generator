from flask import Flask, render_template, request, jsonify, send_file
import io
import zipfile
import openpyxl
from openpyxl.styles import Alignment
from copy import copy
import os
from datetime import datetime, timedelta, date, time
from dateutil.relativedelta import relativedelta

app = Flask(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "Factuur format.xlsx")

MAANDEN_NL = {
    1: "januari", 2: "februari", 3: "maart", 4: "april",
    5: "mei", 6: "juni", 7: "juli", 8: "augustus",
    9: "september", 10: "oktober", 11: "november", 12: "december"
}


def datum_nl(d):
    return f"{d.day} {MAANDEN_NL[d.month]} {d.year}"


def periode_str(data):
    """Geeft de periodestring van één urenregistratie."""
    e = datetime.strptime(data["eerste_datum"], "%Y-%m-%d")
    l = datetime.strptime(data["laatste_datum"], "%Y-%m-%d")
    return datum_nl(e) if e == l else f"{datum_nl(e)} t/m {datum_nl(l)}"


def lees_urenregistratie(bestand_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(bestand_bytes), data_only=True)
    ws = wb.active

    locatie = str(ws.cell(row=2, column=4).value or "")

    naam = ""
    t17 = ws.cell(row=17, column=20).value
    if t17 and str(t17).startswith("Costs "):
        naam = str(t17)[6:].strip()
    if not naam:
        for rij in range(6, 30):
            val = ws.cell(row=rij, column=5).value
            if val:
                naam = str(val)
                break

    werk_uren = reis_uren = wacht_uren = totaal_km = 0.0
    datums = []
    activiteiten = []
    bekende_types = {"worktime", "traveltime", "waitworktime", "worktimeminus lunch"}

    for rij in range(6, 200):
        type_val = ws.cell(row=rij, column=3).value
        if not type_val or not isinstance(type_val, str):
            if rij > 25:
                break
            continue
        if type_val.lower() not in bekende_types:
            break

        datum_val = ws.cell(row=rij, column=2).value
        datum_obj = None
        if datum_val:
            if isinstance(datum_val, datetime):
                datums.append(datum_val)
                datum_obj = datum_val
            elif isinstance(datum_val, date):
                dt = datetime(datum_val.year, datum_val.month, datum_val.day)
                datums.append(dt)
                datum_obj = dt

        van = ws.cell(row=rij, column=6).value
        tot = ws.cell(row=rij, column=7).value
        omschrijving_cel = str(ws.cell(row=rij, column=4).value or "")
        uren = 0.0
        if van is not None and tot is not None:
            if isinstance(van, time) and isinstance(tot, time):
                uren = (tot.hour + tot.minute / 60) - (van.hour + van.minute / 60)
            elif isinstance(van, (int, float)) and isinstance(tot, (int, float)):
                uren = (tot - van) * 24

        datum_str = datum_obj.strftime("%Y-%m-%d") if datum_obj else None
        tl = type_val.lower()
        if tl == "worktimeminus lunch":
            werk_uren -= uren
            for act in reversed(activiteiten):
                if act["type"] == "werk":
                    act["uren"] -= uren
                    break
        elif tl.startswith("worktime"):
            werk_uren += uren
            activiteiten.append({"type": "werk", "omschrijving": omschrijving_cel, "uren": uren, "datum": datum_str})
        elif tl == "traveltime":
            reis_uren += uren
            km = ws.cell(row=rij, column=9).value
            km_val = float(km) if km and isinstance(km, (int, float)) else 0.0
            totaal_km += km_val
            activiteiten.append({"type": "reis", "omschrijving": omschrijving_cel, "uren": uren, "datum": datum_str, "km": km_val})
        elif tl == "waitworktime":
            wacht_uren += uren
            activiteiten.append({"type": "wacht", "omschrijving": omschrijving_cel, "uren": uren, "datum": datum_str})

    # Lees extra kosten op basis van het label in kolom T (20), niet op vaste rij
    # zodat oude én nieuwe uren-registratiebestanden allebei werken
    lunch = bonnetjes = overnachting = 0.0
    for _r in range(16, 23):
        _label = str(ws.cell(row=_r, column=20).value or "").lower()
        _val   = float(ws.cell(row=_r, column=22).value or 0)
        if "lunch"        in _label: lunch        = _val
        elif "bonnetjes"  in _label: bonnetjes    = _val
        elif "overnachting" in _label: overnachting = _val

    # Detecteer of dit bestand met eigen auto gereden is (geen "Auto van:" notitie)
    eigen_auto = True
    for rij in ws.iter_rows():
        for cel in rij:
            if isinstance(cel.value, str) and cel.value.startswith("Auto van:"):
                eigen_auto = False
                break
        if not eigen_auto:
            break

    nu = datetime.now()
    eerste  = min(datums) if datums else nu
    laatste = max(datums) if datums else nu

    return {
        "naam":          naam,
        "locatie":       locatie,
        "eerste_datum":  eerste.strftime("%Y-%m-%d"),
        "laatste_datum": laatste.strftime("%Y-%m-%d"),
        "werk_uren":     round(werk_uren, 10),
        "reis_uren":     round(reis_uren, 10),
        "wacht_uren":    round(wacht_uren, 10),
        "km":            round(totaal_km, 2),
        "eigen_auto":    eigen_auto,
        "lunch":         lunch,
        "bonnetjes":     bonnetjes,
        "overnachting":  overnachting,
        "activiteiten":  activiteiten,
    }


def maak_factuur(uren_data_lijst, client_naam, client_adres, client_postcode,
                 client_email, client_kvk, factuurnummer, btw_pct, eigen_auto, btw_verrekenen=True,
                 factuurdatum=None, vervaldatum=None):

    # Aggregeer alle urenregistraties
    totalen = {
        "naam":       uren_data_lijst[0]["naam"],
        "locatie":    uren_data_lijst[0]["locatie"],
        "werk_uren":  round(sum(d["werk_uren"]  for d in uren_data_lijst), 2),
        "reis_uren":  round(sum(d["reis_uren"]  for d in uren_data_lijst), 2),
        "wacht_uren": round(sum(d["wacht_uren"] for d in uren_data_lijst), 2),
        "km":         round(sum(d["km"]         for d in uren_data_lijst), 2),
        "lunch":      sum(d["lunch"]        for d in uren_data_lijst),
        "bonnetjes":  sum(d["bonnetjes"]    for d in uren_data_lijst),
        "overnachting": sum(d["overnachting"] for d in uren_data_lijst),
    }

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # Factuuradres
    ws.cell(row=8,  column=1, value=client_naam)
    ws.cell(row=9,  column=1, value=client_adres)
    ws.cell(row=10, column=1, value=client_postcode)
    ws.cell(row=11, column=1, value=client_email)
    ws.cell(row=12, column=1, value=client_kvk)

    # Factuurnummer en datums
    try:
        factuur_dt = datetime.strptime(factuurdatum, "%Y-%m-%d") if factuurdatum else datetime.now()
    except (ValueError, TypeError):
        factuur_dt = datetime.now()
    try:
        verval_dt = datetime.strptime(vervaldatum, "%Y-%m-%d") if vervaldatum else factuur_dt + relativedelta(months=1)
    except (ValueError, TypeError):
        verval_dt = factuur_dt + relativedelta(months=1)

    ws.cell(row=15, column=3,  value=factuurnummer)
    ws.cell(row=15, column=12, value=factuurnummer)
    ws.cell(row=16, column=3,  value=factuur_dt)
    ws.cell(row=16, column=12, value="=C16")
    ws.cell(row=17, column=3,  value=verval_dt)
    ws.cell(row=17, column=12, value="=C17")

    # Kolommen C en L breed genoeg voor lange datums ("10 augustus 2026")
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["L"].width = 18
    # Kolommen G/H/I smaller zodat het factuur op twee pagina's past bij afdrukken
    ws.column_dimensions["G"].width = 7
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 11
    ws.column_dimensions["R"].width = 11

    # Periode — één regel per urenregistratie
    periodes = "\n".join(periode_str(d) for d in uren_data_lijst)
    cel = ws.cell(row=13, column=8, value=periodes)
    cel.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[13].height = max(15, len(uren_data_lijst) * 15)

    EERSTE_REG     = 21   # eerste activiteitrij in de template
    TEMPLATE_RIJEN = 15   # template heeft rijen 21–35
    SUBTOTAAL_RIJ  = 36   # subtotaal staat in de originele template op rij 36
    AFBEELDING_RIJ = 47   # blauwe balk (image-in-cell) in de originele template

    # Bewaar rijhoogte, merged-cell-patronen en celopmaak van een template-activiteitrij
    _rh = ws.row_dimensions[EERSTE_REG].height
    template_rij_hoogte = _rh if _rh else (ws.sheet_format.defaultRowHeight or 15)
    template_stijlen = {}
    for col in range(1, 10):  # A t/m I — G/H/I meenemen voor randen en opmaak
        src = ws.cell(row=EERSTE_REG, column=col)
        if src.has_style:
            template_stijlen[col] = {
                "border":        copy(src.border),
                "fill":          copy(src.fill),
                "font":          copy(src.font),
                "alignment":     copy(src.alignment),
                "number_format": src.number_format,
            }

    # Wis alle regelrijen (A, G, H) zodat lege regels volledig leeg zijn
    for r in range(EERSTE_REG, EERSTE_REG + TEMPLATE_RIJEN):
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=7, value=None)
        ws.cell(row=r, column=8, value=None)

    # ── Pass 1: bouw de exacte lijst van alle te schrijven activiteitrijen ──
    rijen = []

    for d in uren_data_lijst:
        e = datetime.strptime(d["eerste_datum"], "%Y-%m-%d")
        l = datetime.strptime(d["laatste_datum"], "%Y-%m-%d")
        dlabel = datum_nl(e) if e == l else f"{datum_nl(e)} t/m {datum_nl(l)}"
        loc = d["locatie"]
        km_d = d["km"] if (d.get("eigen_auto", True) and eigen_auto) else 0

        activiteiten = d.get("activiteiten") or []
        if activiteiten:
            for act in activiteiten:
                if not act["uren"]:
                    continue
                if act.get("datum"):
                    act_dt = datetime.strptime(act["datum"], "%Y-%m-%d")
                    act_datum = datum_nl(act_dt)
                else:
                    act_datum = dlabel
                omschr = act.get("omschrijving") or loc
                if act["type"] == "werk":
                    rijen.append({"omschrijving": f"Gewerkte uren - {act_datum} - {omschr}",
                                  "aantal": act["uren"], "tarief": 45.0, "is_uren": True})
                elif act["type"] == "reis":
                    rijen.append({"omschrijving": f"Vergoeding reisuren - {act_datum} - {omschr}",
                                  "aantal": act["uren"], "tarief": 22.5, "is_uren": True})
                elif act["type"] == "wacht":
                    rijen.append({"omschrijving": f"WachtWerkTijd uren - {act_datum} - {omschr}",
                                  "aantal": act["uren"], "tarief": 22.5, "is_uren": True})
            if km_d:
                rijen.append({"omschrijving": f"Vergoeding KM's - {dlabel} - {loc}",
                              "aantal": km_d, "tarief": 0.35, "is_uren": False})
        else:
            for omschrijving, aantal, tarief, check in [
                (f"Gewerkte uren - {dlabel} - {loc}",       d["werk_uren"],  45.0, d["werk_uren"]),
                (f"Vergoeding reisuren - {dlabel} - {loc}",  d["reis_uren"],  22.5, d["reis_uren"]),
                (f"Vergoeding KM's - {dlabel} - {loc}",      km_d,            0.35, km_d),
                (f"WachtWerkTijd uren - {dlabel} - {loc}",   d["wacht_uren"], 22.5, d["wacht_uren"]),
            ]:
                if check:
                    rijen.append({"omschrijving": omschrijving, "aantal": aantal,
                                  "tarief": tarief, "is_uren": False})

    for label, sleutel in [
        ("(Vergoeding) Lunch",        "lunch"),
        ("(Vergoeding) Overnachting", "overnachting"),
    ]:
        bedragen = [d[sleutel] for d in uren_data_lijst if d[sleutel] > 0]
        if not bedragen:
            continue
        alle_gelijk = len(set(bedragen)) == 1
        rijen.append({"omschrijving": label,
                      "aantal": len(bedragen) if alle_gelijk else 1,
                      "tarief": bedragen[0] if alle_gelijk else sum(bedragen),
                      "is_uren": False})

    for d in uren_data_lijst:
        if d["bonnetjes"] <= 0:
            continue
        e = datetime.strptime(d["eerste_datum"], "%Y-%m-%d")
        l = datetime.strptime(d["laatste_datum"], "%Y-%m-%d")
        datum_label = datum_nl(e) if e == l else f"{datum_nl(e)} t/m {datum_nl(l)}"
        rijen.append({"omschrijving": f"(Vergoeding) Bonnetjes - {datum_label}",
                      "aantal": 1, "tarief": d["bonnetjes"], "is_uren": False})

    # ── Voeg precies het juiste aantal extra rijen in (op basis van werkelijke telling) ──
    extra_rijen = max(0, len(rijen) - TEMPLATE_RIJEN)
    if extra_rijen > 0:
        ws.insert_rows(SUBTOTAAL_RIJ, amount=extra_rijen)

        # openpyxl laat "ghost merges" achter op de originele template-rij-posities
        # (bv. SUBTOTAAL rijen 36-39 en blauwe-balk rijen 47-48 hadden brede A:I merges).
        # Verwijder alle merged-cell-ranges die in het ingevoegde bereik vallen.
        for mr in list(ws.merged_cells.ranges):
            if SUBTOTAAL_RIJ <= mr.min_row < SUBTOTAAL_RIJ + extra_rijen:
                ws.merged_cells.remove(mr)

        for r in range(EERSTE_REG + TEMPLATE_RIJEN, EERSTE_REG + TEMPLATE_RIJEN + extra_rijen):
            ws.row_dimensions[r].height = template_rij_hoogte
            for col, stijl in template_stijlen.items():
                tgt = ws.cell(row=r, column=col)
                tgt.border        = copy(stijl["border"])
                tgt.fill          = copy(stijl["fill"])
                tgt.font          = copy(stijl["font"])
                tgt.alignment     = copy(stijl["alignment"])
                tgt.number_format = stijl["number_format"]

    subtotaal_rij = SUBTOTAAL_RIJ + extra_rijen
    btw_pct_rij   = subtotaal_rij + 1
    btw_euro_rij  = subtotaal_rij + 2
    totaal_rij    = subtotaal_rij + 3

    # Kolommen G en H samenvoegen en centreren voor de totaalrijen
    for rij_nr in [subtotaal_rij, btw_pct_rij, btw_euro_rij, totaal_rij]:
        # Verwijder eventuele bestaande G/H merge op deze rij
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == rij_nr and mr.max_row == rij_nr and mr.min_col <= 8 and mr.max_col >= 7:
                ws.merged_cells.remove(mr)
        ws.merge_cells(start_row=rij_nr, start_column=7, end_row=rij_nr, end_column=8)
        ws.cell(row=rij_nr, column=7).alignment = Alignment(horizontal='center', vertical='center')

    # ── Pass 2: schrijf rijen naar het werkblad ──
    vrije_rij = EERSTE_REG
    for rij in rijen:
        ws.cell(row=vrije_rij, column=1, value=rij["omschrijving"])
        ws.cell(row=vrije_rij, column=7, value=rij["aantal"]).number_format = '0.00'
        ws.cell(row=vrije_rij, column=8, value=rij["tarief"]).number_format = '0.00'
        vrije_rij += 1

    # Ingevoegde rijen missen de =G*H bedragformule uit de template — schrijf die expliciet
    for r in range(EERSTE_REG + TEMPLATE_RIJEN, vrije_rij):
        c = ws.cell(row=r, column=9, value=f"=ROUND(G{r}*H{r},2)")
        c.number_format = '0.00'

    # Zet 2 decimalen op alle activiteitrijen (ook template-rijen 21-35)
    for r in range(EERSTE_REG, vrije_rij):
        ws.cell(row=r, column=7).number_format = '0.00'
        ws.cell(row=r, column=8).number_format = '0.00'
        ws.cell(row=r, column=9).number_format = '0.00'

    # Subtotaal SUM-formule over alle gebruikte activiteitrijen
    ws.cell(row=subtotaal_rij, column=9,
            value=f"=SUM(I{EERSTE_REG}:I{vrije_rij - 1})")

    # BTW percentage en BTW-nummer
    # Q5 altijd leegmaken: template-formule geeft anders 0
    ws.cell(row=5, column=17, value="")
    if not btw_verrekenen:
        ws.cell(row=5, column=8, value="")
        ws.cell(row=btw_pct_rij, column=9, value=0)
    else:
        ws.cell(row=btw_pct_rij, column=9, value=btw_pct / 100)

    # BTW-bedrag en totaal expliciet schrijven zodat ze altijd naar de juiste rijen verwijzen
    ws.cell(row=btw_euro_rij, column=9,
            value=f"=I{subtotaal_rij}*I{btw_pct_rij}")
    ws.cell(row=totaal_rij, column=9,
            value=f"=I{subtotaal_rij}+I{btw_euro_rij}")

    return wb, subtotaal_rij


def herstel_afbeeldingen(template_path, output_buf, subtotaal_rij=36):
    """
    openpyxl strips <drawing>, printerSettings en vm="1" (image-in-cell).
    Deze functie patcht de openpyxl-output zodat alle afbeeldingen behouden blijven.
    """
    import re

    output_buf.seek(0)
    nieuwe_buf = io.BytesIO()

    # Bestanden vanuit openpyxl (gegevens die we hebben aangepast)
    # sheet1.xml.rels komt van het TEMPLATE (behoud rId2 → drawing, rId1 → printerSettings)
    uit_output = {
        'xl/worksheets/sheet1.xml',
        'xl/sharedStrings.xml',
        'xl/styles.xml',              # openpyxl's stijlindices moeten overeenkomen met sheet1.xml
        'xl/workbook.xml',            # bevat fullCalcOnLoad="1" zodat formules herberekend worden
        'xl/_rels/workbook.xml.rels', # bijbehorende relaties van workbook
    }

    # Deze bestanden altijd vanuit template (richData / afbeeldingen)
    altijd_template = {
        'xl/richData/richValueRel.xml',
        'xl/richData/_rels/richValueRel.xml.rels',
        'xl/richData/rdRichValueTypes.xml',
        'xl/richData/rdrichvalue.xml',
        'xl/richData/rdrichvaluestructure.xml',
        'xl/metadata.xml',
        'xl/media/image1.png',
    }

    with zipfile.ZipFile(output_buf, 'r') as out_zip, \
         zipfile.ZipFile(template_path, 'r') as tmpl_zip, \
         zipfile.ZipFile(nieuwe_buf, 'w', zipfile.ZIP_DEFLATED) as nieuw_zip:

        tmpl_namen = set(tmpl_zip.namelist())

        for naam in tmpl_namen:
            if naam == '[Content_Types].xml':
                continue  # Wordt apart afgehandeld onderaan

            elif naam in altijd_template:
                nieuw_zip.writestr(naam, tmpl_zip.read(naam))

            elif naam in uit_output and naam in out_zip.namelist():
                data = out_zip.read(naam)

                if naam == 'xl/worksheets/sheet1.xml':
                    tekst = data.decode('utf-8')

                    # 1. Voeg vm="1" terug toe aan de afbeelding-in-cel rijen
                    # Blauwe balk: 4 rijen onder TOTAAL (TOTAAL = subtotaal_rij + 3)
                    _afb_rij = subtotaal_rij + 7
                    for cel in [f'A{_afb_rij}', f'J{_afb_rij}']:
                        tekst = re.sub(
                            rf'(<c\b[^>]*\br="{cel}"[^>]*?)(\s*>)',
                            lambda m: (m.group(1) + ' vm="1"' + m.group(2))
                                      if 'vm=' not in m.group(0) else m.group(0),
                            tekst
                        )

                    # 2. Drawing element vervangen/toevoegen met rId2
                    # (template-rels: rId2 = drawing1.xml, rId1 = printerSettings)
                    # openpyxl kan al een <drawing r:id="rId1"/> bevatten met zijn eigen rId — verwijder dat
                    tekst = re.sub(r'<drawing\b.*?/>', '', tekst)
                    tekst = tekst.replace(
                        '</worksheet>',
                        '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId2"/></worksheet>'
                    )

                    data = tekst.encode('utf-8')

                elif naam == 'xl/_rels/workbook.xml.rels':
                    # openpyxl's workbook.xml.rels mist de richData-relaties — voeg ze toe
                    tekst = data.decode('utf-8')
                    extra_rels = [
                        ('rId_meta',  'http://schemas.openxmlformats.org/officeDocument/2006/relationships/sheetMetadata',       'metadata.xml'),
                        ('rId_rvrel', 'http://schemas.microsoft.com/office/2022/10/relationships/richValueRel',                  'richData/richValueRel.xml'),
                        ('rId_rv',    'http://schemas.microsoft.com/office/2017/06/relationships/rdRichValue',                   'richData/rdrichvalue.xml'),
                        ('rId_rvs',   'http://schemas.microsoft.com/office/2017/06/relationships/rdRichValueStructure',          'richData/rdrichvaluestructure.xml'),
                        ('rId_rvt',   'http://schemas.microsoft.com/office/2017/06/relationships/rdRichValueTypes',              'richData/rdRichValueTypes.xml'),
                    ]
                    for rid, rtype, target in extra_rels:
                        if target not in tekst:
                            tekst = tekst.replace(
                                '</Relationships>',
                                f'<Relationship Id="{rid}" Type="{rtype}" Target="{target}"/></Relationships>'
                            )
                    data = tekst.encode('utf-8')

                elif naam == 'xl/workbook.xml':
                    # Zorg dat de template-calcChain-referentie ook in openpyxl's workbook.xml zit
                    tekst = data.decode('utf-8')
                    if 'calcChain' not in tekst:
                        # Voeg calcChain relatie toe aan workbook.xml.rels later via rId_calc
                        pass
                    data = tekst.encode('utf-8')

                nieuw_zip.writestr(naam, data)

            else:
                nieuw_zip.writestr(naam, tmpl_zip.read(naam))

        # [Content_Types].xml: gebruik template-versie (bevat al alle richData-types)
        # maar voeg eventuele extra openpyxl-types toe die er nog niet in zitten
        tmpl_ct = tmpl_zip.read('[Content_Types].xml').decode()
        out_ct  = out_zip.read('[Content_Types].xml').decode()
        # Voeg openpyxl-types toe die niet in template zitten
        for override in re.findall(r'<Override[^>]*/>', out_ct):
            partname = re.search(r'PartName="([^"]+)"', override)
            if partname and partname.group(1) not in tmpl_ct:
                tmpl_ct = tmpl_ct.replace('</Types>', override + '</Types>')
        nieuw_zip.writestr('[Content_Types].xml', tmpl_ct)

        # Voeg openpyxl-bestanden toe die niet in het template zitten
        for naam in out_zip.namelist():
            if naam not in tmpl_namen and naam not in uit_output and naam not in altijd_template and naam != '[Content_Types].xml':
                nieuw_zip.writestr(naam, out_zip.read(naam))

    nieuwe_buf.seek(0)
    return nieuwe_buf


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/healthcheck")
def healthcheck():
    import os
    tmpl_ok = os.path.exists(TEMPLATE_PATH)
    tmpl_size = os.path.getsize(TEMPLATE_PATH) if tmpl_ok else 0
    return {
        "template_gevonden": tmpl_ok,
        "template_grootte_bytes": tmpl_size,
        "template_pad": TEMPLATE_PATH,
    }


@app.route("/parse", methods=["POST"])
def parse():
    bestand = request.files.get("bestand")
    if not bestand:
        return jsonify({"succes": False, "fout": "Geen bestand ontvangen."}), 400
    try:
        data = lees_urenregistratie(bestand.read())
        return jsonify({"succes": True, "data": data})
    except Exception as e:
        return jsonify({"succes": False, "fout": str(e)}), 500


@app.route("/genereer", methods=["POST"])
def genereer():
    form = request.get_json(force=True, silent=True) or {}
    uren_data_lijst = form.get("uren_data_lijst", [])
    if not uren_data_lijst:
        return jsonify({"succes": False, "fout": "Geen urendata ontvangen."}), 400

    try:
        wb, sub_rij = maak_factuur(
            uren_data_lijst = uren_data_lijst,
            client_naam     = form.get("client_naam", ""),
            client_adres    = form.get("client_adres", ""),
            client_postcode = form.get("client_postcode", ""),
            client_email    = form.get("client_email", ""),
            client_kvk      = form.get("client_kvk", ""),
            factuurnummer   = form.get("factuurnummer", datetime.now().strftime("%Y%m%d")),
            btw_pct         = float(form.get("btw_pct", 21)),
            eigen_auto      = bool(form.get("eigen_auto", True)),
            btw_verrekenen  = bool(form.get("btw_verrekenen", True)),
            factuurdatum    = form.get("factuurdatum"),
            vervaldatum     = form.get("vervaldatum"),
        )
        buf = io.BytesIO()
        wb.save(buf)
        buf = herstel_afbeeldingen(TEMPLATE_PATH, buf, subtotaal_rij=sub_rij)

        naam = uren_data_lijst[0].get("naam", "")
        fnr  = form.get("factuurnummer", datetime.now().strftime("%Y%m%d"))
        bestandsnaam = f"{fnr} - Factuur {naam}.xlsx".strip()

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=bestandsnaam,
        )
    except Exception as e:
        return jsonify({"succes": False, "fout": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    app.run(debug=True, port=port, host="0.0.0.0")
